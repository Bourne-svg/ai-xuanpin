"""
Excel导出工具
"""
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def to_excel(products: list, stats: dict, video_name: str = "") -> bytes:
    """
    将选品结果导出为Excel文件，返回bytes
    包含两个sheet：产品清单 + 统计概览
    """
    if not HAS_OPENPYXL:
        raise ImportError("请安装openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ============ Sheet 1: 产品清单 ============
    ws1 = wb.active
    ws1.title = "产品清单"

    # 表头样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["排名", "产品名称", "类别", "市场需求度", "差异化空间", "物流可行性",
               "综合评级", "出现次数", "时间位置", "选品建议"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 评级颜色
    rating_fills = {
        "高": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "中": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "低": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for i, p in enumerate(products, 1):
        row = i + 1
        data = [
            i,
            p.get("产品名称", ""),
            p.get("类别", ""),
            p.get("市场需求度", ""),
            p.get("差异化空间", ""),
            p.get("物流可行性", ""),
            p.get("综合评级", ""),
            p.get("出现次数", p.get("出现频率", 1)),
            ", ".join(p.get("时间戳列表", [p.get("时间戳", "")])),
            p.get("选品建议", ""),
        ]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col == 7 and val in rating_fills:
                cell.fill = rating_fills[val]
            if col in (4, 5, 6, 8):
                cell.alignment = Alignment(horizontal="center")

    # 列宽
    col_widths = [6, 28, 12, 10, 10, 10, 10, 10, 25, 40]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ============ Sheet 2: 统计概览 ============
    ws2 = wb.create_sheet("统计概览")

    title_font = Font(bold=True, size=14)
    ws2.cell(row=1, column=1, value=f"AI选品分析报告").font = title_font
    ws2.cell(row=2, column=1, value=f"视频: {video_name}")
    ws2.cell(row=3, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    ws2.cell(row=5, column=1, value="总体统计").font = Font(bold=True, size=12)
    stat_headers = ["指标", "数量"]
    for col, h in enumerate(stat_headers, 1):
        cell = ws2.cell(row=6, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    stat_rows = [
        ("产品总数", stats.get("总数", 0)),
        ("高潜力产品", stats.get("高潜力", 0)),
        ("中潜力产品", stats.get("中潜力", 0)),
        ("低潜力产品", stats.get("低潜力", 0)),
    ]
    for i, (label, val) in enumerate(stat_rows, 7):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=val)

    # 类别分布
    cat_start = len(stat_rows) + 8
    ws2.cell(row=cat_start, column=1, value="类别分布").font = Font(bold=True, size=12)
    ws2.cell(row=cat_start + 1, column=1, value="类别").font = header_font
    ws2.cell(row=cat_start + 1, column=1).fill = header_fill
    ws2.cell(row=cat_start + 1, column=2, value="数量").font = header_font
    ws2.cell(row=cat_start + 1, column=2).fill = header_fill
    for i, (cat, count) in enumerate(stats.get("类别分布", {}).items(), cat_start + 2):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=count)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
